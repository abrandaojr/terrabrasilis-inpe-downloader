"""
06_export_tables.py
===================
Exports all PRODES statistics to a publication-ready Excel workbook.

All PRODES data is queried on-the-fly from the GeoParquet files produced
by script 02.  Reference data (MapBiomas, GFW/FAO) is included as clearly
labelled constant tables.

Excel design: The Economist style
  – Navy / dark-green header rows
  – Alternating off-white row fill
  – Thin cell borders, no gridlines beyond the table
  – Bold totals and highlights
  – Source notes below every table
  – Two workbooks: PT-BR and EN-US

Output
------
  C:\\Amintas\\Prodes\\tables\\PRODES_Statistics_PT_<date>.xlsx
  C:\\Amintas\\Prodes\\tables\\PRODES_Statistics_EN_<date>.xlsx

Usage
-----
    python 06_export_tables.py

Author
------
Amintas Brandão Jr. <abrandaojr@gmail.com>
Imazon — Instituto do Homem e Meio Ambiente da Amazônia

License
-------
MIT
"""

from __future__ import annotations

__version__ = "1.0.0"
__all__: list[str] = []

import importlib.util
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

HERE        = Path(__file__).parent
TABLES_DIR  = Path(r"C:\Amintas\Prodes\tables")
GPQ_DIR     = Path(r"C:\Amintas\Prodes\geoparquet")

# ---------------------------------------------------------------------------
# Dependency bootstrap
# ---------------------------------------------------------------------------

def _bootstrap(*packages: tuple[str, str]) -> None:
    import importlib, shutil
    mod_by_pip = {pip: mod for pip, mod in packages}

    def _still_missing(pkgs: list[str]) -> list[str]:
        importlib.invalidate_caches()
        return [p for p in pkgs if not importlib.util.find_spec(mod_by_pip[p])]

    missing = _still_missing(list(mod_by_pip))
    if not missing:
        return
    if not shutil.which("uv"):
        subprocess.call([sys.executable, "-m", "pip", "install", "--quiet", "uv"],
                        stderr=subprocess.DEVNULL)
    for base in [
        [sys.executable, "-m", "pip", "install", "--quiet"],
        ["uv", "pip", "install", "--python", sys.executable, "--quiet"],
        [sys.executable, "-m", "pip", "install", "--quiet", "--break-system-packages"],
    ]:
        if not missing:
            return
        try:
            subprocess.check_call(base + missing, stderr=subprocess.DEVNULL)
            missing = _still_missing(missing)
        except (subprocess.CalledProcessError, FileNotFoundError):
            pass
    if missing:
        sys.exit(f"[FATAL] Could not install: {' '.join(missing)}")


_bootstrap(
    ("duckdb",    "duckdb"),
    ("pyarrow",   "pyarrow"),
    ("openpyxl",  "openpyxl"),
)

import duckdb                          # noqa: E402
import pyarrow.parquet as pq           # noqa: E402
from openpyxl import Workbook          # noqa: E402
from openpyxl.styles import (          # noqa: E402
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter  # noqa: E402

SEP = "=" * 65
DIV = "-" * 65

# ---------------------------------------------------------------------------
# Excel palette (The Economist / NYT hybrid)
# ---------------------------------------------------------------------------

_H_DARK   = "1B3A4B"   # dark navy   — header background
_H_MID    = "2E7D32"   # forest green — accent headers
_H_FONT   = "FFFFFF"   # white        — header text
_ROW_ALT  = "F4F6F8"   # very light gray — alternating rows
_ROW_WHT  = "FFFFFF"   # white        — default rows
_TXT_DARK = "111111"   # near-black   — body text
_TXT_MED  = "555555"   # medium gray  — secondary text
_TXT_RED  = "C0392B"   # warning red  — negative / high values
_TXT_GRN  = "2E7D32"   # forest green — positive / low values
_BRD_CLR  = "CCCCCC"   # thin border

_THIN  = Side(style="thin",   color=_BRD_CLR)
_THICK = Side(style="medium", color="888888")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BORDER_BOTTOM_THICK = Border(left=_THIN, right=_THIN, top=_THIN,
                               bottom=Side(style="medium", color="333333"))

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color=_TXT_DARK, italic=False) -> Font:
    return Font(name="Calibri", bold=bold, size=size,
                color=color, italic=italic)

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ---------------------------------------------------------------------------
# Reference data (non-PRODES sources)
# ---------------------------------------------------------------------------

_COVER_PCT = {
    "pt": [
        ("Pantanal",       86.0),
        ("Amazônia",       81.0),
        ("Caatinga",       60.0),
        ("Cerrado",        53.0),
        ("Pampa",          38.0),
        ("Mata Atlântica", 13.0),
    ],
    "en": [
        ("Pantanal",         86.0),
        ("Amazon",           81.0),
        ("Caatinga",         60.0),
        ("Cerrado Savanna",  53.0),
        ("Pampa",            38.0),
        ("Atlantic Forest",  13.0),
    ],
}

_INTL = [
    ("Brazil",      9_064, "INPE/PRODES (national methodology)"),
    ("D.R. Congo",  4_900, "GFW primary forest loss"),
    ("Bolivia",     4_200, "GFW primary forest loss"),
    ("Indonesia",   2_800, "GFW primary forest loss"),
    ("Colombia",    1_450, "GFW primary forest loss"),
]

_POLICY_TARGETS = {2026: 4_866, 2028: 4_000}

# ---------------------------------------------------------------------------
# DuckDB helpers (reused from script 04)
# ---------------------------------------------------------------------------

_BIOME_TO_LABEL_PT = {
    "Amazon Biome": "Amazônia Legal",
    "Legal Amazon": "Amazônia Legal",
    "Cerrado":      "Cerrado",
    "Caatinga":     "Caatinga",
    "Pantanal":     "Pantanal",
    "Mata Atlantica": "Mata Atlântica",
    "Pampa":        "Pampa",
}
_BIOME_TO_LABEL_EN = {
    "Amazon Biome": "Legal Amazon",
    "Legal Amazon": "Legal Amazon",
    "Cerrado":      "Cerrado Savanna",
    "Caatinga":     "Caatinga",
    "Pantanal":     "Pantanal",
    "Mata Atlantica": "Atlantic Forest",
    "Pampa":        "Pampa",
}
_DEFOR_KW  = ("deforestation", "desmatamento", "desmat")
_AUX_KW    = ("border", "boundary", "hydrography", "indigenous",
               "conservation_units", "settlement", "quilombola")
_AREA_COLS = ("areakm", "area_km", "area_km2", "area")
_YEAR_COLS = ("year", "ano", "yr")
_AMAZON_DIRS = {"Amazon Biome", "Legal Amazon"}
_BIOME_NAMES = set(_BIOME_TO_LABEL_PT)


def _discover_defor_files(gpq_dir: Path) -> dict[str, list[Path]]:
    biome_files: dict[str, list[Path]] = {}
    for pf in sorted(gpq_dir.rglob("*.parquet")):
        try:
            parts    = pf.relative_to(gpq_dir).parts
            path_low = "/".join(p.lower() for p in parts)
            if not any(k in path_low for k in _DEFOR_KW):
                continue
            if any(k in path_low for k in _AUX_KW):
                continue
            biome = next((p for p in parts if p in _BIOME_NAMES), None)
            if biome:
                biome_files.setdefault(biome, []).append(pf)
        except (ValueError, IndexError):
            pass
    return biome_files


def _detect_cols(files: list[Path]) -> tuple[str | None, str | None]:
    for f in files[:5]:
        try:
            schema = pq.read_schema(str(f))
            low_to_orig = {n.lower(): n for n in schema.names}
            ac = next((low_to_orig[c] for c in _AREA_COLS if c in low_to_orig), None)
            yc = next((low_to_orig[c] for c in _YEAR_COLS if c in low_to_orig), None)
            if ac:
                return ac, yc
        except Exception:
            pass
    return None, None


def _infer_factor(files: list[Path], area_col: str) -> float:
    sample: list[float] = []
    for f in files[:3]:
        try:
            col = pq.read_table(str(f), columns=[area_col]).column(area_col).to_pylist()
            sample += [v for v in col if v and v > 0][:20]
        except Exception:
            pass
    if not sample:
        return 1.0
    med = sorted(sample)[len(sample) // 2]
    return 1 / 1_000_000 if med > 500_000 else (1 / 100 if med > 5_000 else 1.0)


def _series_by_year(files: list[Path], ac: str, yc: str, factor: float) -> list[tuple[int, float]]:
    paths = [str(f).replace("\\", "/") for f in files]
    sql = f"""
        SELECT CAST("{yc}" AS INTEGER) AS yr,
               SUM(CAST("{ac}" AS DOUBLE)) * {factor} AS km2
        FROM   read_parquet({paths!r})
        WHERE  "{yc}" IS NOT NULL AND "{ac}" IS NOT NULL AND CAST("{ac}" AS DOUBLE) > 0
        GROUP  BY yr HAVING yr BETWEEN 2000 AND 2030
        ORDER  BY yr
    """
    try:
        rows = duckdb.connect().execute(sql).fetchall()
        return [(int(r[0]), round(float(r[1]), 1)) for r in rows if r[0] and r[1]]
    except Exception:
        return []


def _series_by_state(files: list[Path], ac: str, yc: str, factor: float,
                     year: int) -> list[tuple[str, float]]:
    paths = [str(f).replace("\\", "/") for f in files]
    state_col = None
    for f in files[:3]:
        try:
            low_map = {n.lower(): n for n in pq.read_schema(str(f)).names}
            state_col = next(
                (low_map[c] for c in ("estado", "state", "uf", "sigla_uf") if c in low_map),
                None
            )
            if state_col:
                break
        except Exception:
            pass
    if not state_col:
        return []
    sql = f"""
        SELECT "{state_col}" AS st,
               SUM(CAST("{ac}" AS DOUBLE)) * {factor} AS km2
        FROM   read_parquet({paths!r})
        WHERE  CAST("{yc}" AS INTEGER) = {year}
          AND  "{ac}" IS NOT NULL AND CAST("{ac}" AS DOUBLE) > 0
          AND  "{state_col}" IS NOT NULL
        GROUP  BY st ORDER BY km2 DESC
    """
    try:
        rows = duckdb.connect().execute(sql).fetchall()
        return [(str(r[0]), round(float(r[1]), 1)) for r in rows if r[0] and r[1]]
    except Exception:
        return []


# ---------------------------------------------------------------------------
# Excel cell writers
# ---------------------------------------------------------------------------

def _write_header_row(ws, row: int, cols: list[str], col_start: int = 1,
                      color: str = _H_DARK) -> None:
    for j, val in enumerate(cols, col_start):
        cell = ws.cell(row=row, column=j, value=val)
        cell.font      = _font(bold=True, color=_H_FONT, size=10)
        cell.fill      = _fill(color)
        cell.alignment = _align("center")
        cell.border    = Border(bottom=Side(style="medium", color="FFFFFF"),
                                right=Side(style="thin", color="FFFFFF"))


def _write_data_row(ws, row: int, values: list, col_start: int = 1,
                    highlight: str | None = None) -> None:
    bg = _ROW_ALT if (row % 2 == 0) else _ROW_WHT
    for j, val in enumerate(values, col_start):
        cell = ws.cell(row=row, column=j, value=val)
        cell.fill      = _fill(bg)
        cell.border    = _BORDER
        cell.alignment = _align("right" if isinstance(val, (int, float)) else "left")
        f_color = _TXT_DARK
        if highlight and isinstance(val, (int, float)):
            f_color = highlight
        cell.font = _font(color=f_color)


def _write_source(ws, row: int, text: str, col_span: int = 6) -> None:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font      = _font(italic=True, size=8, color=_TXT_MED)
    cell.alignment = _align("left")


def _write_section_title(ws, row: int, title: str, col_span: int = 6) -> None:
    cell = ws.cell(row=row, column=1, value=title)
    cell.font      = _font(bold=True, size=12, color=_H_DARK)
    cell.alignment = _align("left")


def _set_col_widths(ws, widths: dict[int, float]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _freeze(ws, cell: str = "B2") -> None:
    ws.freeze_panes = cell


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _sheet_amazon_series(wb: Workbook, lang: str,
                         series: list[tuple[int, float]]) -> None:
    name = "Série Histórica" if lang == "pt" else "Historical Series"
    ws   = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    _freeze(ws, "B3")

    cols_pt = ["Ano", "Desmatamento (km²)", "Variação Anual (%)",
               "% da Meta 2028", "Meta (km²)"]
    cols_en = ["Year", "Deforestation (km²)", "Annual Change (%)",
               "% of 2028 Target", "Target (km²)"]
    cols = cols_pt if lang == "pt" else cols_en

    title = ("Desmatamento Anual na Amazônia Legal · INPE/PRODES"
             if lang == "pt" else
             "Annual Deforestation in Brazil's Legal Amazon · INPE/PRODES")
    _write_section_title(ws, 1, title)
    _write_header_row(ws, 2, cols)

    target_2028 = 4_000
    for i, (yr, km2) in enumerate(series, 3):
        prev_km2 = series[i - 4][1] if i > 3 else None
        yoy      = round((km2 - prev_km2) / prev_km2 * 100, 1) if prev_km2 else None
        pct_tgt  = round(km2 / target_2028 * 100, 1)
        tgt      = _POLICY_TARGETS.get(yr, "")
        vals     = [yr, round(km2, 0), yoy, pct_tgt, tgt]
        _write_data_row(ws, i, vals,
                        highlight=_TXT_RED if km2 > target_2028 * 2 else None)

    # Targets
    for yr, km2 in _POLICY_TARGETS.items():
        i = len(series) + 3
        _write_data_row(ws, i, [yr, km2, "", 100.0, km2])
        i += 1

    row_src = len(series) + len(_POLICY_TARGETS) + 4
    src = ("Fonte: INPE/PRODES. Calculado a partir dos dados GeoParquet."
           if lang == "pt" else
           "Source: INPE/PRODES. Calculated from GeoParquet data.")
    _write_source(ws, row_src, src)

    _set_col_widths(ws, {1: 8, 2: 22, 3: 20, 4: 18, 5: 16})


def _sheet_biome_comparison(wb: Workbook, lang: str,
                             biome_data: dict[str, tuple[float, int]]) -> None:
    name = "Por Bioma" if lang == "pt" else "By Biome"
    ws   = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    _freeze(ws, "B3")

    cols_pt = ["Bioma", "Desmatamento (km²)", "Ano de Referência",
               "% do Total Brasileiro"]
    cols_en = ["Biome",  "Deforestation (km²)", "Reference Year",
               "% of Brazilian Total"]
    cols = cols_pt if lang == "pt" else cols_en

    title = ("Desmatamento por Bioma Brasileiro · INPE/PRODES"
             if lang == "pt" else
             "Deforestation by Brazilian Biome · INPE/PRODES")
    _write_section_title(ws, 1, title)
    _write_header_row(ws, 2, cols, color=_H_MID)

    total = sum(v for v, _ in biome_data.values()) or 1
    biome_map = _BIOME_TO_LABEL_PT if lang == "pt" else _BIOME_TO_LABEL_EN
    rows = sorted(biome_data.items(), key=lambda x: x[1][0], reverse=True)

    for i, (biome_key, (km2, yr)) in enumerate(rows, 3):
        label = biome_map.get(biome_key, biome_key)
        pct   = round(km2 / total * 100, 1)
        _write_data_row(ws, i, [label, round(km2, 0), yr, pct])

    # Totals row
    total_row = len(rows) + 3
    cell = ws.cell(row=total_row, column=1, value="TOTAL" if lang == "en" else "TOTAL")
    cell.font = _font(bold=True)
    cell = ws.cell(row=total_row, column=2, value=round(total, 0))
    cell.font = _font(bold=True)
    ws.cell(row=total_row, column=4, value=100.0).font = _font(bold=True)

    src = ("Fonte: INPE/PRODES. Calculado a partir dos dados GeoParquet."
           if lang == "pt" else
           "Source: INPE/PRODES. Calculated from GeoParquet data.")
    _write_source(ws, total_row + 2, src)
    _set_col_widths(ws, {1: 22, 2: 22, 3: 18, 4: 22})


def _sheet_state_breakdown(wb: Workbook, lang: str,
                            state_data: list[tuple[str, float]],
                            year: int) -> None:
    name = f"Por Estado ({year})" if lang == "pt" else f"By State ({year})"
    ws   = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False

    cols_pt = ["Estado", "Desmatamento (km²)", "% do Total Amazônia"]
    cols_en = ["State",  "Deforestation (km²)", "% of Amazon Total"]
    cols = cols_pt if lang == "pt" else cols_en

    title = (f"Desmatamento na Amazônia Legal por Estado · {year} · INPE/PRODES"
             if lang == "pt" else
             f"Legal Amazon Deforestation by State · {year} · INPE/PRODES")
    _write_section_title(ws, 1, title)
    _write_header_row(ws, 2, cols, color=_H_MID)

    total = sum(v for _, v in state_data) or 1
    for i, (state, km2) in enumerate(state_data, 3):
        pct = round(km2 / total * 100, 1)
        _write_data_row(ws, i, [state, round(km2, 0), pct])

    src = ("Fonte: INPE/PRODES. Calculado a partir dos dados GeoParquet."
           if lang == "pt" else
           "Source: INPE/PRODES. Calculated from GeoParquet data.")
    _write_source(ws, len(state_data) + 4, src)
    _set_col_widths(ws, {1: 28, 2: 22, 3: 22})


def _sheet_forest_cover(wb: Workbook, lang: str) -> None:
    name = "Cobertura Florestal" if lang == "pt" else "Forest Cover"
    ws   = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False

    cols_pt = ["Bioma", "Cobertura Remanescente (%)", "Avaliação"]
    cols_en = ["Biome",  "Remaining Cover (%)",        "Status"]
    cols = cols_pt if lang == "pt" else cols_en

    title = ("Cobertura de Vegetação Nativa Remanescente · MapBiomas 2023"
             if lang == "pt" else
             "Remaining Native Vegetation Cover · MapBiomas 2023")
    _write_section_title(ws, 1, title)
    _write_header_row(ws, 2, cols, color=_H_MID)

    ratings_pt = {(0, 30): "Crítico", (30, 60): "Preocupante",
                  (60, 80): "Moderado", (80, 100): "Preservado"}
    ratings_en = {(0, 30): "Critical", (30, 60): "Concerning",
                  (60, 80): "Moderate", (80, 100): "Preserved"}
    ratings = ratings_pt if lang == "pt" else ratings_en

    data = _COVER_PCT[lang]
    for i, (biome, pct) in enumerate(sorted(data, key=lambda x: x[1]), 3):
        rating = next(v for (lo, hi), v in ratings.items() if lo <= pct < hi)
        row_vals = [biome, pct, rating]
        _write_data_row(ws, i, row_vals)
        color = _TXT_RED if pct < 40 else (_TXT_MED if pct < 65 else _TXT_GRN)
        ws.cell(row=i, column=2).font = _font(bold=True, color=color)

    src = ("Fonte: MapBiomas 2023. Mapeamento Anual da Cobertura e Uso da Terra no Brasil."
           if lang == "pt" else
           "Source: MapBiomas 2023. Annual Land Cover and Use Mapping Project.")
    _write_source(ws, len(data) + 4, src)
    _set_col_widths(ws, {1: 22, 2: 26, 3: 18})


def _sheet_international(wb: Workbook, lang: str) -> None:
    name = "Comparativo Global" if lang == "pt" else "International"
    ws   = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False

    cols_pt = ["País", "Perda Floresta Tropical (km²)", "Fonte / Metodologia"]
    cols_en = ["Country", "Tropical Forest Loss (km²)", "Source / Methodology"]
    cols = cols_pt if lang == "pt" else cols_en

    title = ("Perda de Floresta Tropical por País · 2023 · GFW / FAO"
             if lang == "pt" else
             "Tropical Forest Loss by Country · 2023 · GFW / FAO")
    _write_section_title(ws, 1, title)
    _write_header_row(ws, 2, cols)

    for i, (country, km2, source) in enumerate(
            sorted(_INTL, key=lambda x: x[1], reverse=True), 3):
        _write_data_row(ws, i, [country, km2, source])
        if country == "Brazil":
            for j in range(1, 4):
                ws.cell(row=i, column=j).font = _font(bold=True, color=_H_MID)

    src = ("* Metodologias distintas: GFW mede perda de cobertura arbórea; "
           "PRODES mede desmatamento.\n  Comparação indicativa, não direta."
           if lang == "pt" else
           "* Different methodologies: GFW measures tree-cover loss; "
           "PRODES measures deforestation.\n  Indicative comparison only.")
    _write_source(ws, len(_INTL) + 4, src)
    _set_col_widths(ws, {1: 18, 2: 28, 3: 38})


def _sheet_methodology(wb: Workbook, lang: str) -> None:
    name = "Metodologia" if lang == "pt" else "Methodology"
    ws   = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False

    notes = {
        "pt": [
            ("INPE/PRODES",
             "Sistema de Monitoramento do Desmatamento na Amazônia Brasileira por Satélite. "
             "Detecta desmatamento por corte raso em imagem óptica. "
             "Referência: https://terrabrasilis.dpi.inpe.br"),
            ("MapBiomas 2023",
             "Projeto de mapeamento anual da cobertura e uso da terra no Brasil. "
             "Dados de cobertura florestal remanescente. "
             "Referência: https://mapbiomas.org"),
            ("Global Forest Watch / FAO",
             "Plataforma global de monitoramento florestal. "
             "Dados de perda de cobertura arbórea (Hansen et al.). "
             "Metodologia diferente do PRODES — não diretamente comparável. "
             "Referência: https://globalforestwatch.org"),
            ("Metas de Desmatamento",
             "Meta de 4.000 km² para 2028 baseada em compromissos do governo brasileiro "
             "e financiamento NORAD/Fundo Amazônia. "
             "Meta intermediária de 4.866 km² para 2026."),
            ("Cálculos",
             "Todos os dados PRODES calculados on-the-fly a partir dos arquivos "
             "GeoParquet gerados pelo pipeline de conversão (script 02). "
             "Gerado em: " + datetime.now().strftime("%Y-%m-%d %H:%M")),
        ],
        "en": [
            ("INPE/PRODES",
             "Brazil's Amazon Deforestation Monitoring System by Satellite. "
             "Detects clear-cut deforestation via optical imagery. "
             "Reference: https://terrabrasilis.dpi.inpe.br"),
            ("MapBiomas 2023",
             "Annual land cover and use mapping project for Brazil. "
             "Remaining forest cover data. "
             "Reference: https://mapbiomas.org"),
            ("Global Forest Watch / FAO",
             "Global forest monitoring platform. "
             "Tree-cover loss data (Hansen et al.). "
             "Different methodology from PRODES — not directly comparable. "
             "Reference: https://globalforestwatch.org"),
            ("Deforestation Targets",
             "4,000 km² target for 2028 based on Brazilian government commitments "
             "and NORAD/Amazon Fund financing. "
             "Intermediate target of 4,866 km² for 2026."),
            ("Calculations",
             "All PRODES data computed on-the-fly from GeoParquet files "
             "generated by the conversion pipeline (script 02). "
             "Generated: " + datetime.now().strftime("%Y-%m-%d %H:%M")),
        ],
    }[lang]

    _write_section_title(ws, 1, "Notas Metodológicas" if lang == "pt" else "Methodological Notes")
    row = 3
    for source, note in notes:
        ws.cell(row=row,   column=1, value=source).font = _font(bold=True, color=_H_DARK, size=11)
        ws.cell(row=row+1, column=1, value=note).font   = _font(color=_TXT_DARK, size=10)
        ws.cell(row=row+1, column=1).alignment = _align(wrap=True)
        ws.row_dimensions[row+1].height = 45
        row += 3

    ws.column_dimensions["A"].width = 90


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    print(f"\n{SEP}")
    print(f"  PRODES Table Exporter  v{__version__}  |  {now}")
    print(f"{SEP}\n")

    if not GPQ_DIR.exists():
        sys.exit(f"[FATAL] GeoParquet directory not found: {GPQ_DIR}\n"
                 "        Run  python 02_convert_to_geoparquet.py  first.")

    TABLES_DIR.mkdir(parents=True, exist_ok=True)

    # ── Load data ────────────────────────────────────────────────────────
    print("  Loading data from GeoParquet files...")
    biome_files = _discover_defor_files(GPQ_DIR)

    if not biome_files:
        sys.exit("[FATAL] No deforestation parquet files found.\n"
                 "        Run  python 02_convert_to_geoparquet.py  first.")

    # Amazon series
    amazon_files = [f for bd, fs in biome_files.items()
                    if bd in _AMAZON_DIRS for f in fs]
    ac, yc = _detect_cols(amazon_files)
    factor = _infer_factor(amazon_files, ac) if ac else 1.0
    amazon_series: list[tuple[int, float]] = []
    if ac and yc:
        amazon_series = _series_by_year(amazon_files, ac, yc, factor)
        print(f"  Amazon series: {len(amazon_series)} year(s) loaded")

    # Biome totals (most recent year per biome)
    ref_year = max(y for y, _ in amazon_series) if amazon_series else None
    biome_data: dict[str, tuple[float, int]] = {}   # biome → (km², year)
    for biome_dir, files in biome_files.items():
        ac2, yc2 = _detect_cols(files)
        if not ac2:
            continue
        fac2 = _infer_factor(files, ac2)
        yr_to_use = ref_year
        if yc2 and ref_year:
            try:
                paths = [str(f).replace("\\", "/") for f in files]
                row = duckdb.connect().execute(
                    f'SELECT MAX(CAST("{yc2}" AS INTEGER)) FROM read_parquet({paths!r})'
                ).fetchone()
                if row and row[0]:
                    yr_to_use = int(row[0])
            except Exception:
                pass
        if yc2 and yr_to_use:
            paths = [str(f).replace("\\", "/") for f in files]
            try:
                row = duckdb.connect().execute(f"""
                    SELECT SUM(CAST("{ac2}" AS DOUBLE)) * {fac2}
                    FROM read_parquet({paths!r})
                    WHERE CAST("{yc2}" AS INTEGER) = {yr_to_use}
                    AND "{ac2}" IS NOT NULL
                """).fetchone()
                if row and row[0] and float(row[0]) > 0:
                    biome_data[biome_dir] = (round(float(row[0]), 1), yr_to_use)
            except Exception:
                pass

    print(f"  Biome data: {list(biome_data)}")

    # State breakdown (Amazon, most recent year)
    state_data: list[tuple[str, float]] = []
    if amazon_files and ac and yc and ref_year:
        state_data = _series_by_state(amazon_files, ac, yc, factor, ref_year)
        print(f"  State breakdown: {len(state_data)} state(s) for {ref_year}")

    # ── Build workbooks ─────────────────────────────────────────────────
    date_str = datetime.now().strftime("%Y-%m-%d")

    for lang in ("pt", "en"):
        wb = Workbook()
        wb.remove(wb.active)   # remove default sheet

        _sheet_amazon_series(wb, lang, amazon_series)
        _sheet_biome_comparison(wb, lang, biome_data)
        if state_data and ref_year:
            _sheet_state_breakdown(wb, lang, state_data, ref_year)
        _sheet_forest_cover(wb, lang)
        _sheet_international(wb, lang)
        _sheet_methodology(wb, lang)

        lang_label = "PT" if lang == "pt" else "EN"
        out = TABLES_DIR / f"PRODES_Statistics_{lang_label}_{date_str}.xlsx"
        wb.save(str(out))
        print(f"\n  Saved [{lang_label}]: {out}")

    print(f"\n{DIV}")
    print(f"  Output folder: {TABLES_DIR}")
    print(f"{SEP}\n")


if __name__ == "__main__":
    main()
